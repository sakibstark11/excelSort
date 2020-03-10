import sys
import openpyxl
from openpyxl.reader.excel import load_workbook
from openpyxl import Workbook
import os
from os import path

def run(file_location):
    file_root = file_location.replace('.xlsx', '/')
    main_book = openpyxl.load_workbook(file_location)
    main_sheet = main_book.active
    print(main_sheet)
    current_book = ''
    current_sheet = ''
    current_file_name = ''
    new_file = False
    data = ''
    for row in main_sheet.iter_rows():
        data = ()
        if row[1].value == None:
            print('New Sheet')
            try:
                if new_file:
                    print('Saving New File ' + file_root + current_file_name + '.xlsx')
                    current_book.save(str(file_root+current_file_name+'.xlsx'))
                current_file_name = str(row[0].value).replace(':', '_').replace(' ', '_')
                current_book = Workbook()
                current_sheet = current_book.active
                new_file = True
            except Exception as err:
                print(err)
        for cell in row:
            if (cell.value != None):
                data = data + (str(cell.value),)
        current_sheet.append(data)

if (len(sys.argv) < 2):
    print('Source file full directory required, if on Linux use linux formart, if on Windows use windows format')
    sys.exit()
else:
    location = str(sys.argv[1]).replace("\\", "/")
    print('Location: ' + location)
    if path.exists(location):
        file_root = location.replace('.xlsx', '/')	
        os.mkdir(file_root)
        run(location)
    else:
        print('Source location does not exist, Ensure correct directory and file extension like .xlsx')
